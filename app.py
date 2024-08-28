from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, redirect, session
from wtforms import StringField, SubmitField, IntegerField
from sqlalchemy import create_engine, text
from flask_wtf import FlaskForm
import pandas as pd
import openpyxl
from openpyxl import Workbook
from flask_session import Session
import PIL
import qrcode
from qrcode.image.styledpil import StyledPilImage
from qrcode.image.styles.moduledrawers.pil import RoundedModuleDrawer
from qrcode.image.styles.colormasks import RadialGradiantColorMask
import datetime
import os
import pythoncom
import win32com.client
import math
import json
from dotenv import load_dotenv


# TODO: Learn what a logger is. (python logging module)


def initialize_com():
    pythoncom.CoInitialize()


#  Create engine
df = pd.read_excel("D:/Work/POOF/complete_product_list_spreadsheet_updated.xlsx")
current_quotation = []
current_client = {
    "Date": "",
    "Customer_Name": "",
    "Customer_Number": "",
    "Rep_Name": "",
    "Rep_Number": "",
}


# Load the environment variables
load_dotenv()
username = os.getenv("USER")
password = os.getenv("PASSWORD")
schema = os.getenv("SCHEMA")

quotation_dir = "static/quotations"
excel_quotation_dir = "static/editable_quotations"
if not os.path.exists(quotation_dir):
    os.mkdir(quotation_dir)


engine = create_engine(
    f"mysql+mysqlconnector://{username}:{password}@127.0.0.1/{schema}"
)
#df.to_sql(name="product_list", con=engine, index=False)


try:
    df.to_sql(name="product_list", con=engine, if_exists="fail", index=False)
except ValueError:
    pass


conn = engine.connect()

# Keep track of the quotation ID
no_of_quotations = conn.execute(
    text("SELECT MAX(quotation_id) FROM exported_quotations")
).all()[0][0]
if not no_of_quotations:
    no_of_quotations = 0
quotation_id = no_of_quotations + 1

# Connect Flask
app = Flask(__name__)

# Configure Session to use filesystem
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# TODO: Add a function that allows user to display the status of the 10 most recent quotations and allow them to download the PDF file
# This function should only allow the user to download the PDF file if the status is approved

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "GET":
        # If the user is not logged in, redirect to login page
        if not session.get("name"):
            return redirect("/login")
        name = session["name"].split(" ")[0]
        # Clear the current quotation and client info
        # Get the 10 most recent quotations
        # Schema: quotation_id, submission_date, employee_id, quotation_url, quotation_file_path, status
        """
            <th scope = "col">Quotation ID</th>
            <th scope = "col">Submission Date</th>
            <th scope = "col">Submitted By</th>
            <th scope = "col">Current Status</th>
            <th scope = "col">Action</th>
        """
        quotations = conn.execute(
            text("SELECT * FROM exported_quotations ORDER BY submission_date DESC LIMIT 10")
        ).all()
        conn.commit()
        new_quotations = []
        # Create a new list that fits the schema of the frontend for ease of use
        for quotation in quotations:
            # Get the employee name
            employee = conn.execute(text(f"SELECT Employee_Name FROM authorized_personnel WHERE employee_id = {quotation[2]}")).all()
            conn.commit()
            employee = employee[0]
            # Add quotation to list of quotations (list of lists)
            pdf_path = os.path.join(quotation_dir, f"{quotation[0]}.pdf")
            # If the status is approved, add who it's approved by
            approved_by = None
            if quotation[5] == "Approved":
                approved_by = conn.execute(text(f"SELECT approved_by FROM exported_quotations WHERE quotation_id = {quotation[0]}")).all()
                conn.commit()
                approved_by = approved_by[0][0]
            # If the status is approved, add the quotation to the list of new quotations
            new_quotations.append([quotation[0], quotation[1], employee[0], quotation[5], quotation[3], pdf_path, approved_by])
        user = session["access_level"]
        if user == "Developer" or user == "Administrator":
            user = "Administrator"
        # Pass available options 
        # Get all admins from the database that have already submit
        admins = conn.execute(text("SELECT Employee_Name FROM authorized_personnel WHERE Authority_Level = 'Administrator' OR Authority_level = 'Developer'")).all()
        conn.commit()
        # Get all users from the database
        users = conn.execute(text("SELECT Employee_Name FROM authorized_personnel")).all()
        conn.commit()

        return render_template("index.html", quotations=new_quotations, user=user, name=name, admins=admins, users=users)
    elif request.method == "POST":
        return redirect("/review")
    # Create a page that displays the status of the 10 most recent quotations and allows the user to download the PDF file if the status is approved

@app.route("/search", methods=["GET", "POST"])
def search():
    quotation_id = request.form.get("quotation_id")
    from_date = request.form.get("from_date")
    to_date = request.form.get("to_date")
    employee_name = request.form.get("user_option")
    current_status = request.form.get("current_status")
    approved_by = request.form.get("approved_by")
    name = session["name"].split(" ")[0]
    user = session["access_level"]
    user = session["access_level"]
    if user == "Developer" or user == "Administrator":
        user = "Administrator"
    sets = []
    # Rewrite the Search function to accomodate the date range and the multiple choice options
    
    if quotation_id:
        quotation_id = set(conn.execute(text(f"SELECT * FROM exported_quotations WHERE quotation_id = {quotation_id}")).all())
        conn.commit()
    if submission_date:
        submission_date = conn.execute(text(f"SELECT * FROM exported_quotations WHERE submission_date LIKE '%{submission_date}'%")).all()
        conn.commit()
        submission_date = set(submission_date)
        sets.append(submission_date)
    if employee_name:
        employee_name = conn.execute(text(f"""SELECT 
                                                exported_quotations.quotation_id,
                                                exported_quotations.submission_date,
                                                exported_quotations.quotation_url,
                                                exported_quotations.quotation_file_path,
                                                authorized_personnel.Employee_Name,
                                                authorized_personnel.Authority_Level
                                            FROM 
                                                poof_schema.exported_quotations
                                            JOIN 
                                                poof_schema.authorized_personnel
                                            ON 
                                                exported_quotations.employee_id = authorized_personnel.employee_id
                                            WHERE 
                                                authorized_personnel.Employee_Name LIKE '%{employee_name}%'""")).all()
        conn.commit()
        employee_name = set(employee_name)
        sets.append(employee_name)
    if current_status:
        current_status = conn.execute(text(f"SELECT * FROM exported_quotations WHERE status = '{current_status}'")).all()
        conn.commit()
        current_status = set(current_status)
        sets.append(current_status)
    if approved_by:
        approved_by = conn.execute(text(f"SELECT * FROM exported_quotations WHERE approved_by LIKE '%{approved_by}%'")).all()
        conn.commit()
        approved_by = set(approved_by)
        sets.append(approved_by)

    print(sets)
    non_empties = []
    for x in sets:
        if x == set():
            continue
        sets.append(x)
    if non_empties:
        result = set.intersection(*non_empties)
    if quotation_id:
        result = quotation_id
    new_quotations = []
    # Create a new list that fits the schema of the frontend for ease of use
    for quotation in result:
        # Get the employee name
        employee = conn.execute(text(f"SELECT Employee_Name FROM authorized_personnel WHERE employee_id = {quotation[2]}")).all()
        conn.commit()
        employee = employee[0]
        # Add quotation to list of quotations (list of lists)
        pdf_path = os.path.join(quotation_dir, f"{quotation[0]}.pdf")
        # If the status is approved, add who it's approved by
        approved_by = None
        if quotation[5] == "Approved":
            approved_by = conn.execute(text(f"SELECT approved_by FROM exported_quotations WHERE quotation_id = {quotation[0]}")).all()
            conn.commit()
            approved_by = approved_by[0][0]
        # If the status is approved, add the quotation to the list of new quotations
        new_quotations.append([quotation[0], quotation[1], employee[0], quotation[5], quotation[3], pdf_path, approved_by])
    
    return render_template("search_results.html", quotations=new_quotations, user=user, name=name)

@app.route("/review", methods=["GET", "POST"])
def review_quotation():
    if request.method == "POST":
        quotation_id = request.form.get("quotation_id")
        # Read the excel and put the values into the entries variable
        # Get the path of the quotation file
        path = os.path.join(excel_quotation_dir, f"{quotation_id}.xlsx")
        path = os.path.abspath(path)
        # Open the quotation file
        requested_quotation = openpyxl.load_workbook(path)
        sheet = requested_quotation.active
        # Assign the values of the cells to the client_data dictionary
        Date_Cell = sheet["I5"]
        Customer_Name_Cell = sheet["A9"]
        Customer_Number_Cell = sheet["C9"]
        Rep_Name_Cell = sheet["A12"]
        Rep_Number_Cell = sheet["C12"]
        Quotation_Number_Cell = sheet["G5"]

        # Create a dictionary to store the client data
        client_data = {
            "Date": None,
            "Customer_Name": None,
            "Customer_Number": None,
            "Rep_Name": None,
            "Rep_Number": None,
        }

        # Assign the values of the cells to the client_data dictionary
        client_data["Date"] = Date_Cell.value
        client_data["Customer_Name"] = Customer_Name_Cell.value
        client_data["Customer_Number"] = Customer_Number_Cell.value
        client_data["Rep_Name"] = Rep_Name_Cell.value
        client_data["Rep_Number"] = Rep_Number_Cell.value

        entries = []
        # Get the values of the cells and put them into the entries list
        # Repeat this for all sheets in the quotation and skip empty lines. Don't look for an image directory if there's no price (aka it's a spec)
        # Get number of sheets in the quotation and iterate over them
        num_sheets = len(requested_quotation.sheetnames)
        for i in range(num_sheets):
            sheet = requested_quotation.worksheets[i]
            rows = sheet.iter_rows(min_row=14, max_row=21, min_col=1, max_col=10)
            for i, row in enumerate(rows):
                if not row[0].value and not row[9].value:
                    break
                quantity = row[6].value
                description = row[0].value
                if row[0].value == None:
                    description = row[9].value
                
                price = row[7].value
                total = row[8].value
                if description is None:
                    continue    
                elif price is None:
                    continue
                else:
                    dir = conn.execute(text(f"SELECT Image_Directory FROM product_list WHERE product_name = '{description}' OR Description = '{description}'"))
                    directories = dir.all()
                    conn.commit()
                    entries.append(
                        [f"{directories[0][0]}", description, price, quantity, total]
                    )
        total = 0
        vat = 0
        for entry in entries:
            if entry[2] is None:
                continue
            total += int(float(entry[4]))
        vat = total * 0.14
        total += vat
        quotation_pdf = os.path.join(quotation_dir, f"{quotation_id}.pdf")
        return render_template("review_quotation.html", entries=entries, total=total, vat=vat, quotation_pdf=quotation_pdf, quotation_id=quotation_id, customer_info=client_data)

@app.route("/approve", methods=["GET", "POST"])
def approve():
    if request.method == "POST":
        print(f"Quotation ID: {request.form.get('quotation_id')}")
        table_data = json.loads(request.form.get("table_data"))
        # Delete the quotation file from the editable_quotations directory and the PDF file from the quotations directory, and the QR code
        os.remove(os.path.join(excel_quotation_dir, f"{request.form.get('quotation_id')}.xlsx"))
        os.remove(os.path.join(quotation_dir, f"{request.form.get('quotation_id')}.pdf"))
        os.remove(os.path.join(quotation_dir, f"{request.form.get('quotation_id')}.png"))
        # Put the data in its proper form in product data and client data and clear the current quotation and client info then call the submit function
        # Clear the current quotation and client info
        current_quotation.clear()
        print(f"Current Client: {current_client}")  
        # Get the data from the table_data and put it in the current_quotation and current_client dictionaries
        """
        current_quotation.append(
            [
                prod[0][0], # Product Code NOT
                prod[0][1], # Product Name NOT
                prod[0][5], # Image DONE
                prod[0][3], # Description DONE
                prod[0][4], # Specs NOT
                quantity,   # Quantity DONE
                prod[0][2], # Price DONE
                prod[0][2] * quantity # Total DONE
            ]
        )
        """

        for data in table_data:
            image = data["image"]
            description = data["description"]
            price = data["price"]
            quantity = data["quantity"]
            sum = data["sum"]
            remaining_data = conn.execute(text(f"SELECT product_code, product_name, Specs FROM product_list WHERE Description = '{description}'")).all()
            conn.commit()
            product_code, product_name, specs = remaining_data[0][0], remaining_data[0][1], remaining_data[0][2]
            current_quotation.append(
                [
                    product_code, # Product Code
                    product_name, # Product Name
                    image, # Image
                    description, # Description
                    specs, # Specs
                    quantity, # Quantity
                    price, # Price
                    sum # Total
                ]
            )
        current_client["Date"] = request.form.get("date")
        current_client["Customer_Name"] = request.form.get("customer_name")
        current_client["Customer_Number"] = request.form.get("customer_number")
        current_client["Rep_Name"] = request.form.get("rep_name")
        current_client["Rep_Number"] = request.form.get("rep_number")

        # Update the status of the quotation to approved
        conn.execute(text(f"UPDATE exported_quotations SET status = 'Approved' WHERE quotation_id = {request.form.get('quotation_id')}"))
        # Update approved by to the name of the user
        conn.execute(text(f"UPDATE exported_quotations SET approved_by = '{session['name']}' WHERE quotation_id = {request.form.get('quotation_id')}"))
        conn.commit()
        # Call the submit function to submit the quotation to the database and export it as a PDF but this time with the new data and the new status
        submit(True)
        return redirect("/")

@app.route("/reject", methods=["GET", "POST"])
def reject():
    if request.method == "POST":
        # Delete the quotation file from the editable_quotations directory and the PDF file from the quotations directory, and the QR code
        os.remove(os.path.join(excel_quotation_dir, f"{request.form.get('quotation_id')}.xlsx"))
        os.remove(os.path.join(quotation_dir, f"{request.form.get('quotation_id')}.pdf"))
        os.remove(os.path.join(quotation_dir, f"{request.form.get('quotation_id')}.png"))
        # Update the status of the quotation to rejected
        conn.execute(text(f"UPDATE exported_quotations SET status = 'Rejected' WHERE quotation_id = {request.form.get('quotation_id')}"))
        conn.commit()
        return redirect("/")

@app.route("/login", methods=["GET", "POST"])
def login():

    # If the user is not logged in, redirect to login page
    if request.method == "GET":
        if session.get("name") and session.get("access_level"):
            print(session.get("name"))
            print(session.get("access_level"))
            return render_template("index.html")
        return render_template("signIn.html")

    elif request.method == "POST":
        # Get username and password
        name = request.form.get("Username")
        Pass = request.form.get("Password")
        session["name"] = name
        # Query Database for the data of the employee
        # TODO Protect against SQL injection
        authorityres = conn.execute(
            text(f'SELECT * FROM authorized_personnel WHERE Employee_Name = "{name}"')
        )
        conn.commit()
        authority = authorityres.all()

        # Ensure username and password were submitted
        if not name or not Pass:
            print("Name or Pass is empty")
            return render_template("invalid_credentials.html", name=name, Pass=Pass)

        # Ensure username exists and password is correct
        if len(authority) != 1 or not check_password_hash(authority[0][2], Pass):
            print("Wrong name or wrong Password")
            return render_template("invalid_credentials.html", authority=authorityres)

        session["user_id"] = authority[0][0]
        session["access_level"] = authority[0][3]
        # Redirect user to appropriate page depending on the authority level
        if session["access_level"]:
            return redirect("/")
        return render_template("invalid_credentials.html")


@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":
        # Get username and password and authority level

        name = request.form.get("username")
        password = request.form.get("up")
        cpassword = request.form.get("up2")
        authority = request.form.get("authority")
        hash = generate_password_hash(password, method="pbkdf2", salt_length=16)
        # print(hash)
        if not name:
            return render_template("invalid_credentials.html")

        # Ensure password was submitted
        elif not password:
            return render_template("invalid_credentials.html")

        # Query database for username
        # TODO Protect against SQL injection
        result = conn.execute(
            text(f'SELECT * FROM authorized_personnel WHERE Employee_Name = "{name}"')
        )
        rows = result.all()

        # Ensure username exists and password is correct
        if len(rows) != 0:
            return render_template("invalid_credentials.html")

        if password != cpassword:
            return render_template("invalid_credentials.html")

        # TODO Protect against SQL injection
        conn.execute(
            text(
                f"""INSERT INTO authorized_personnel(Employee_Name, Password, Authority_Level) 
                VALUES ("{name}", "{hash}", "{authority}")"""
            )
        )
        conn.commit()

        # Redirect user to home page
        return redirect("/")

    else:
        return render_template("register.html")


# Logout the user
@app.route("/logout")
def logout():
    # Clear the session
    session.clear()
    return redirect("/")


# Tasks page
@app.route("/task", methods=["GET", "POST"])
def task():
    if request.method == "GET":

        match session["access_level"]:
            case "Developer":
                return render_template("admin_options.html")
            case "Administrator":
                return render_template("admin_options.html")
            case "Data_Entry":
                return render_template("customer_info.html")
            case _:
                return render_template("invalid_credentials.html")

    elif request.method == "POST":
        choice = request.form.get("task")
        # print(choice)
        match choice:
            case "create_quotation":
                return render_template("customer_info.html")
            case "Edit product prices":
                return render_template("edit_prices.html")
            case "Add new product":
                return render_template("add_product.html")
            case "Add Employee":
                return render_template("register.html")
            case _:
                return render_template("invalid_choice.html")


@app.route("/customer_info", methods=["GET", "POST"])
def Customer_Info():
    if request.method == "POST":
        current_client["Date"] = request.form.get("quotation_date")
        current_client["Customer_Name"] = request.form.get("customer_name")
        current_client["Customer_Number"] = request.form.get("customer_number")
        current_client["Rep_Name"] = request.form.get("rep_name")
        current_client["Rep_Number"] = request.form.get("rep_number")

        rows = conn.execute(
            text(
                f"SELECT product_code, product_name FROM product_list ORDER BY product_code ASC"
            )
        )
        prods = rows.all()
        if (
            not current_client["Customer_Name"]
            or not current_client["Date"]
            or not current_client["Rep_Name"]
            or not current_client["Rep_Number"]
        ):
            return render_template("insufficient_data.html")
        return render_template(
            "create_quotation.html", products=prods, customer_info=current_client
        )
    pass


@app.route("/edit_quotation", methods=["GET", "POST"])
def edit_quotation():
    if request.method == "POST":
        rows = conn.execute(
            text(
                f"SELECT product_code, product_name FROM product_list ORDER BY product_code ASC"
            )
        )
        prods = rows.all()
        return render_template(
            "create_quotation.html",
            products=prods,
            customer_info=current_client,
            entries=current_quotation,
        )



# Create Quotation Page and handling Queries, autocomplete, and dynamic table row insertion
@app.route("/create_quotation", methods=["GET", "POST"])
def create_quotation():
    if request.method == "POST":
        # Get the product code and quantity from the submitted form
        code = request.form.get("product_code")
        quantity = request.form.get("quantity")
        if quantity is None:
            return render_template("insufficient_data.html")
        quantity = float(quantity)
        rows = conn.execute(
            text(f"SELECT * FROM product_list WHERE product_code = '{code}'")
        )
        prod = rows.all()
        # redo the sql table and replace it with a list of lists in the frontend instead of a SQL table

        if len(prod) == 0 or prod is None:
            return render_template("insufficient_data.html")

        # Add the product to the current quotation
        current_quotation.append(
            [
                prod[0][0], # Product Code
                prod[0][1], # Product Name
                prod[0][5], # Image
                prod[0][3], # Description
                prod[0][4], # Specs
                quantity,   # Quantity    
                prod[0][2], # Price
                prod[0][2] * quantity # Total
            ]
        )

        rows = conn.execute(
            text(
                f"SELECT product_code, product_name FROM product_list ORDER BY product_code ASC"
            )
        )
        prods = rows.all()
        return render_template(
            "create_quotation.html",
            entries=current_quotation,
            products=prods,
            customer_info=current_client,
        )


# TODO Add a route that allows the user to delete a product from the quotation. This will require a new form for each table row in the HTML


@app.route("/preview", methods=["GET", "POST"])
def preview():
    # Render a preview page that shows the current quotation, an option to go back and edit, and an option to submit
    # Process current_quotation to put it in the following format: [Product_Name, Description, Price, Quantity, Sum]
    entries = []
    for prod in current_quotation:
        # Get the product name, description, price, quantity, and sum
        product_name = prod[1]
        description = prod[3]
        price = prod[6]
        quantity = prod[5]
        sum = prod[7]
        # Get the image directory
        dir = conn.execute(text(f"SELECT Image_Directory FROM product_list WHERE product_name = '{product_name}' OR Description = '{description}'"))
        directories = dir.all()
        conn.commit()
        # Add the product to the entries list
        entries.append(
            [f"{directories[0][0]}", description, price, quantity, sum]
        )

    return render_template(
        "preview_quotation.html",
        customer_info=current_client,
        entries=entries,
        quotation_id = quotation_id
    )

@app.route("/edited_current_quotation", methods=["GET", "POST"])
def edited_current_quotation():
    if request.method == "POST":
        table_data = json.loads(request.form.get("table_data"))
        current_quotation.clear()
        for data in table_data:
            image = data["image"]
            description = data["description"]
            price = data["price"]
            quantity = data["quantity"]
            sum = data["sum"]
            remaining_data = conn.execute(text(f"SELECT product_code, product_name, Specs FROM product_list WHERE Description = '{description}'")).all()
            conn.commit()
            product_code, product_name, specs = remaining_data[0][0], remaining_data[0][1], remaining_data[0][2]
            current_quotation.append(
                [
                    product_code, # Product Code
                    product_name, # Product Name
                    image, # Image
                    description, # Description
                    specs, # Specs
                    quantity, # Quantity
                    price, # Price
                    sum # Total
                ]
            )
        return redirect("/export")


def vacant_spots(sheet):
    vacants = 0
    num_rows_per_sheet = 8
    min_row = 14
    max_row = min_row + num_rows_per_sheet - 1
    min_col = 1
    max_col = 10
    rows = sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    )
    for row in rows:
        if row[0].value is None:
            vacants += 1
    return vacants


def add_small_product_to_page(product_specs, biggest_product, ws, vacancy, product_data):

    # The product has 8 or less lines
    # Check if the product can fit in the current page
    size = product_specs[biggest_product][1] + 1
    if size > vacancy:
        raise Exception("Not enough vacancy for product sent to add_small_product_to_page")
    for product in product_specs:
        if product == biggest_product:
            # Add the product name to the first cell of the page
            description = list(product_data.loc[product_data["Description"] == product]["Description"])[0]
            quantity = list(product_data.loc[product_data["Description"] == product]["Quantity"])[0]
            price = list(product_data.loc[product_data["Description"] == product]["Price"])[0]
            total = list(product_data.loc[product_data["Description"] == product]["Total"])[0]
            product_name = list(product_data.loc[product_data["Description"] == product]["Product_Name"])[0]
            if description == None:
                ws.cell(row=14, column=1).value = product_name
            else:
                ws.cell(row=14, column=1).value = product
            ws.cell(row=14, column=7).value = quantity
            ws.cell(row=14, column=8).value = price
            ws.cell(row=14, column=9).value = total
            ws.cell(row=14, column=10).value = product_name
            for i, spec in enumerate(product_specs[product][0]):
                ws.cell(row=15 + i + 8 - vacancy, column=1).value = spec
            break
    return True


def empty_page(ws):
    # Empty the page
    for row in ws.iter_rows(min_row=14, max_row=21, min_col=1, max_col=10):
        row[0].value = None
        row[6].value = None
        row[7].value = None
        row[8].value = None
        row[9].value = None
    return ws


def add_large_product_to_page(product_specs, biggest_product, ws, num_sheets, sheets, quotation_temp, product_data):
    # If the product has more than 8 lines, split it into multiple page
    # Add the first 8 lines of the product to the first page and remove them from the dictionary
    # Repeat the process for the rest of the pages
    current_page_number = len(sheets)  # The current page number
    # Add number of sheets needed to the sheets list
    for _ in range(num_sheets):
            ws = quotation_temp.copy_worksheet(ws)
            # Empty the page
            empty_page(ws)
            # Add the new page to the sheets list
            sheets.append(ws)

    for i in range(current_page_number, current_page_number + num_sheets, 1):
        ws = sheets[i]
        for product in product_specs:
            if product == biggest_product:
                # Add the product name to the first cell of the page
                # Only add the product name to the first page
                if i == current_page_number:
                    print(f"Product Data:\n {product_data.to_string()}")
                    description = list(product_data.loc[product_data["Description"] == product]["Description"])[0]
                    quantity = list(product_data.loc[product_data["Description"] == product]["Quantity"])[0]
                    price = list(product_data.loc[product_data["Description"] == product]["Price"])[0]
                    total = list(product_data.loc[product_data["Description"] == product]["Total"])[0]
                    product_name = list(product_data.loc[product_data["Description"] == product]["Product_Name"])[0]
                    if description == None:
                        ws.cell(row=14, column=1).value = product_data["Product_Name"][0]
                    else:
                        ws.cell(row=14, column=1).value = product
                    # Get Quantity, Price, Total, and Product Name from the product_data dictionary where product name is equal to product
                    
                    
                    ws.cell(row=14, column=7).value = quantity
                    ws.cell(row=14, column=8).value = price
                    ws.cell(row=14, column=9).value = total
                    ws.cell(row=14, column=10).value = product_name
                    # Add the first 7 lines of the specs to the page
                    for j in range(7):
                        ws.cell(row=15 + j, column=1).value = product_specs[product][0][j]
                    # Remove the first 7 lines from the product
                    product_specs[product][0] = product_specs[product][0][7:]
                    product_specs[product][1] -= 7
                # if it's not the first page, keep adding the specs to the page until the page is full or the product is finished
                else:
                    # If the product has less than 8 lines left, add them to the page
                    if product_specs[product][1] <= 8:
                        for j, spec in enumerate(product_specs[product][0]):
                            ws.cell(row=14 + j, column=1).value = spec
                        # Remove the specs from the product
                        product_specs[product][0] = []
                        product_specs[product][1] = 0
                        break
                    # If the product has more than 8 lines left, add the first 8 lines to the page
                    else:
                        for j in range(8):
                            ws.cell(row=14 + j, column=1).value = product_specs[product][0][j]
                        # Remove the first 8 lines from the product
                        product_specs[product][0] = product_specs[product][0][8:]
                        product_specs[product][1] -= 8
                        break

# Convert the current quotation to a dataframe, submit it to the database, and export it as an excel file
@app.route("/export", methods=["GET", "POST"])
def submit(sent_to_approve=False):
    # TODO: Break this down into atomic functions

    if request.method == "GET":
        pass
    
    client_columns = [
        "Date",
        "Customer_Name",
        "Customer_Number",
        "Rep_Name",
        "Rep_Number",
    ]
    product_columns = [

        "Product_Code",
        "Product_Name",
        "Image",
        "Description",
        "Specs",
        "Quantity",
        "Price",
        "Total",
    ]

    client_list = [
        [
            current_client["Date"],
            current_client["Customer_Name"],
            current_client["Customer_Number"],
            current_client["Rep_Name"],
            current_client["Rep_Number"],
        ]
    ]
    client_data = pd.DataFrame(client_list, columns=client_columns)
    product_data = pd.DataFrame(current_quotation, columns=product_columns)
    # Export the dataframes to an Excel file, then save the file as a pdf
    # and save the pdf to the database alongside the name of the user and the date of submission
    # Add serializtion to the quotation files
    # Clear the current quotation and client info
    current_quotation.clear()
    current_client.clear()
    no_of_quotations = conn.execute(
        text("SELECT MAX(quotation_id) FROM exported_quotations")
    ).all()[0][0]
    if not no_of_quotations:
        no_of_quotations = 0
    if not sent_to_approve:
        quotation_id = no_of_quotations + 1
    else:
        quotation_id = no_of_quotations
    # Open Template file
    quotation_temp = openpyxl.load_workbook("D:/Work/POOF/Quotation_Template.xlsx")
    ws = quotation_temp.active
    sheets = [ws]
    Date_Cell = ws["I5"]
    Customer_Name_Cell = ws["A9"]
    Customer_Number_Cell = ws["C9"]
    Rep_Name_Cell = ws["A12"]
    Rep_Number_Cell = ws["C12"]
    Quotation_Number_Cell = ws["G5"]
    
    # Add the client data to the quotation template
    Date_Cell.value = client_data["Date"][0]
    Customer_Name_Cell.value = client_data["Customer_Name"][0]
    Customer_Number_Cell.value = client_data["Customer_Number"][0]
    Rep_Name_Cell.value = client_data["Rep_Name"][0]
    Rep_Number_Cell.value = client_data["Rep_Number"][0]
    Quotation_Number_Cell.value = quotation_id
    
    
    # Get the true number of rows by unpacking product data from the database
    product_specs = {}
    for code in product_data["Product_Code"]:
        rows = conn.execute(
            text(f"SELECT Description, Specs FROM product_list WHERE product_code = '{code}'")
        )
        specs = rows.all()
        # If there's no specs, set the product specs to None
        product_name = specs[0][0]
        if specs[0][1] == None:
            product_specs[product_name] = None
        # If there are specs, split them and store them in the product_specs dictionary, along with the number of specs as a list
        else:
            # product_specs["Product_Name"][0] is the product name, product_specs["Product_Name"][1] is the number of specs
            product_specs[product_name] = [specs[0][1].split("@"), len(specs[0][1].split("@"))]
    # Sort the product specs by the number of specs in descending order
    product_specs = dict(sorted(product_specs.items(), key=lambda item: item[1][1], reverse=True))
    
    while len(product_specs) > 0:

        # Check current page vacancy and save the number of vacant spots in a variable vacancy
        vacancy = vacant_spots(ws)
        # Check the biggest product in the dictionary and save its size in variable size and its name in variable largest_product
        biggest_product = list(product_specs.keys())[0]
        size = product_specs[biggest_product][1] + 1
        # If size is less than 8:

        if size <= 8:
            # If vacancy is greater than or equal to size, add the product to the current page and remove it from the dictionary
            if vacancy >= size:
                add_small_product_to_page(product_specs, biggest_product, ws, vacancy, product_data)
                del product_specs[biggest_product]
            # Elif vacancy is less than size:
            elif vacancy < size:
                found = False
                # Go through the rest of the products in the dictionary and check if any of them can fit in the current page
                for product in product_specs:
                    temp_size = product_specs[product][1] + 1
                    if temp_size <= vacancy:
                        # If a product is found, add it to the current page and remove it from the dictionary
                        add_small_product_to_page(product_specs, product, ws, vacancy, sheets, quotation_temp, product_data)
                        del product_specs[product]
                        found = True
                        break
                # If no product is found, move to the next page and add the product to the new page
                if not found:
                    ws = quotation_temp.copy_worksheet(ws)
                    sheets.append(ws)
                    # Empty the page
                    empty_page(ws)
        # Elif size is greater than 8:
        elif size > 8:
            # Create as many pages as needed to fit the product
            num_sheets = math.ceil(product_specs[biggest_product][1] / 8)
            # product_specs, biggest_product, ws, num_sheets, sheets, quotation_temp, product_data
            add_large_product_to_page(product_specs, biggest_product, ws, num_sheets, sheets, quotation_temp, product_data)
            del product_specs[biggest_product]
    # Go over the total column in each page and calculate the total for all pages
    total = 0
    for sheet in sheets:
        for row in sheet.iter_rows(min_row=14, max_row=21, min_col=1, max_col=10):
            if row[8].value is None:
                break
            print(f"Row: {row[8].value}")
            total += int(float(row[8].value))
    # Add total to the last page only at cell I23
    ws["I23"].value = total
    

    # If the user is an admin, use the admin_submit function to submit the quotation
    # If the user is not an admin, use the non_admin_submit function to submit the quotation
    print(f"Sent to Approve: {sent_to_approve}")
    if sent_to_approve:
        submit_quotation_to_db(session["user_id"], quotation_temp, ws, True)
    else:
        submit_quotation_to_db(session["user_id"], quotation_temp, ws)

    # Clear Dataframes
    client_data = pd.DataFrame()
    product_data = pd.DataFrame()
    
    # Redirect the user to the successful submission page and redirect after 5 seconds
    return render_template("successful_submission.html")
    

#Create a page that the user is redirected to after submitting a quotation that allows them to download the PDF file
""" @app.route("/download", methods=["GET"])
def download():
    no_of_quotations = conn.execute(
        text("SELECT MAX(quotation_id) FROM exported_quotations")
    ).all()[0][0]
    if not no_of_quotations:
        no_of_quotations = 0
    quotation_id = no_of_quotations
    pdf_path = os.path.join(quotation_dir, f"{quotation_id}.pdf")
    print(f"PDF Path: {pdf_path}")
    return render_template("successful_submission.html", pdf_path=pdf_path) """

# Create Page that allows admins to change the price of products
@app.route("/price", methods=["GET", "POST"])
def price():
    rows = conn.execute(
        text(
            f"SELECT product_code, product_name FROM product_list ORDER BY product_code ASC"
        )
    )
    prods = rows.all()
    if request.method == "GET":
        return render_template("edit_prices.html")

    choice = request.form.get("price")
    if choice == "single_product":
        return render_template("single_product.html", products=prods)
    elif choice == "percentage":
        return render_template("percentage.html")


# Change the price of a single product
@app.route("/single_price", methods=["GET", "POST"])
def single_price():
    rows = conn.execute(
        text(
            f"SELECT product_code, product_name FROM product_list ORDER BY product_code ASC"
        )
    )
    prods = rows.all()
    if request.method == "POST":
        code = request.form.get("product_code")
        price = request.form.get("price")
        conn.execute(
            text(
                f'UPDATE product_list SET Price = "{price}" WHERE Product_Code = "{code}"'
            )
        )
        conn.commit()
        return render_template("index.html")
    return render_template("single_product.html", products=prods)


# Change all product prices by a constant percentage either by increasing or decreasing
@app.route("/percentage", methods=["GET", "POST"])
def percentage():
    if request.method == "POST":
        percentage = request.form.get("percentage")
        Type = request.form.get("change_type")
        if not percentage:
            return render_template("invalid_choice.html")
        if Type == "Increase":
            new_percentage = 1 + (float(percentage) / 100.0)
        elif Type == "Decrease":
            new_percentage = 1 - (float(percentage) / 100.0)
        conn.execute(text(f"UPDATE product_list SET Price = Price * {new_percentage}"))
        conn.commit()
        return render_template("edit_prices.html")
    return render_template("percentage.html")


# Add a new product to the database
@app.route("/add_product", methods=["GET", "POST"])
def add_product():
    if request.method == "POST":
        code = request.form.get("code")
        name = request.form.get("name")
        price = request.form.get("price")
        description = request.form.get("description")
        image = request.form.get("image")
        conn.execute(
            text(
                f'INSERT INTO product_list(Product_Code, Product_Name, Price, Description, Image_Directory) VALUES ("{code}", "{name}", "{price}", "{description}", "{image}")'
            )
        )
        conn.commit()
        return render_template("admin_options.html")
    return render_template("add_product.html")

# View the Quotation
@app.route("/view", methods=["GET"])  # expection a url/view?quotation_id=1
def view_quotation(quotation_id = None):
    if quotation_id is None:
        quotation_id = request.args.get("quotation_id")

    # Read the excel and put the values into the entries variable
    # Get the path of the quotation file
    path = os.path.join(excel_quotation_dir, f"{quotation_id}.xlsx")
    path = os.path.abspath(path)
    # Open the quotation file
    requested_quotation = openpyxl.load_workbook(path)
    sheet = requested_quotation.active
    # Assign the values of the cells to the client_data dictionary
    Date_Cell = sheet["I5"]
    Customer_Name_Cell = sheet["A9"]
    Customer_Number_Cell = sheet["C9"]
    Rep_Name_Cell = sheet["A12"]
    Rep_Number_Cell = sheet["C12"]
    Quotation_Number_Cell = sheet["G5"]

    # Create a dictionary to store the client data
    client_data = {
        "Date": None,
        "Customer_Name": None,
        "Customer_Number": None,
        "Rep_Name": None,
        "Rep_Number": None,
    }

    # Assign the values of the cells to the client_data dictionary
    client_data["Date"] = Date_Cell.value
    client_data["Customer_Name"] = Customer_Name_Cell.value
    client_data["Customer_Number"] = Customer_Number_Cell.value
    client_data["Rep_Name"] = Rep_Name_Cell.value
    client_data["Rep_Number"] = Rep_Number_Cell.value

    entries = []
    # Get the values of the cells and put them into the entries list
    # Repeat this for all sheets in the quotation and skip empty lines. Don't look for an image directory if there's no price (aka it's a spec)
    # Get number of sheets in the quotation and iterate over them
    num_sheets = len(requested_quotation.sheetnames)
    for i in range(num_sheets):
        sheet = requested_quotation.worksheets[i]
        rows = sheet.iter_rows(min_row=14, max_row=21, min_col=1, max_col=10)
        for i, row in enumerate(rows):
            if not row[0].value and not row[9].value:
                break
            quantity = row[6].value
            description = row[0].value
            if row[0].value == None:
                description = row[9].value
            
            price = row[7].value
            total = row[8].value
            if description is None:
                continue    
            elif price is None:
                entries.append(
                    ["bullet.png", description, 0, 0, 0]
                )
            else:
                dir = conn.execute(text(f"SELECT Image_Directory FROM product_list WHERE product_name = '{description}' OR Description = '{description}'"))
                directories = dir.all()
                conn.commit()
                entries.append(
                    [f"{directories[0][0]}", description, price, quantity, total]
                )

    """ 
    # Get the values of the cells and put them into the entries list
    rows = sheet.iter_rows(min_row=14, max_row=21, min_col=1, max_col=10)
    print("Rows: ", rows)
    for i, row in enumerate(rows):
        if not row[0].value and not row[9].value:
            break
        quantity = row[6].value

        # Get the Image_Directory from database
        description = row[0].value 
        if row[0].value == None:
            description = row[9].value
        dir = conn.execute(text(f"SELECT Image_Directory FROM product_list WHERE product_name = '{description}' OR Description = '{description}'"))
        directories = dir.all()
        conn.commit()
        print(f"Description: {description}")
        price = row[7].value
        total = row[8].value
        entries.append(
            [f"{directories[0][0]}", description, price, quantity, total]
        ) """

    total = 0
    vat = 0
    for entry in entries:
        if entry[2] is None:
            continue

        total += int(float(entry[4]))
    vat = total * 0.14
    total += vat
    quotation_pdf = os.path.join(quotation_dir, f"{quotation_id}.pdf")
    return render_template("view_quotation.html", entries=entries, total=total, vat=vat, quotation_pdf=quotation_pdf)


@app.route("/price_list", methods=["GET", "POST"])
def price_list():
    prods = conn.execute(text(f"SELECT product_code, product_name FROM product_list"))
    products = prods.all()
    conn.commit()

    if request.method == "POST":
        code = request.form.get("product_code")
        prods = conn.execute(
            text(
                f"SELECT product_code, product_name, price FROM product_list WHERE product_code = '{code}'"
            )
        )
        data = prods.all()
        if len(data) > 0:
            plist = [[data[0][0], data[0][1], data[0][2]]]
        conn.commit()

        return render_template("price_list.html", plist=plist, products=products)
    return render_template("price_list.html", products=products)

""" # Split the (submit_quotation_to_db) function into 2 smaller functions.
# One function will submit the quotation to the database and save to excel only, this one will be called if the user is not an admin
# The other function will change the quotation's status to approved, remove the existing excel file, replace it with a new one that has the QR code
# and save the new file as a pdf. This one will be called if the user is an admin

# Submit function for non admins and without QR code or PDF
def non_admin_submit(employee_id: int, quotations, sheet) -> bool:
    no_of_quotations = conn.execute(
        text("SELECT MAX(quotation_id) FROM exported_quotations")
    ).all()[0][0]
    if not no_of_quotations:
        no_of_quotations = 0

    quotation_id = no_of_quotations + 1
    submission_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    quotation_url = f"http://localhost:5000/view?quotation_id={quotation_id}"  # TODO: Change this to the actual URL
    quotation_file_path = os.path.join(excel_quotation_dir, f"{quotation_id}.xlsx")
    # Check access level of the user, if the user is an admin, set the status to approved, else set it to pending
    if session["access_level"] == "Administrator" or session["access_level"] == "Developer":
        raise Exception("User is an admin, this function is for non-admins")
    else:
        status = "Pending"
    conn.execute(
        text(
            f"INSERT INTO exported_quotations(submission_date, employee_id, quotation_url, quotation_file_path, status) VALUES ('{submission_date}', {employee_id}, '{quotation_url}', '{quotation_file_path}', '{status}')"
        )
    )
    conn.commit()

    # Save the quotation to the file system
    quotations.save(filename=quotation_file_path)

    return True


def admin_submit(quotation_id: int, quotations, sheet) -> bool:
    # Check if the user is an admin
    if session["access_level"] != "Administrator" and session["access_level"] != "Developer":
        raise Exception("User is not an admin, this function is for admins")
    # Get the quotation file path
    quotation_file_path = os.path.join(excel_quotation_dir, f"{quotation_id}.xlsx")
    pdf_file_path = os.path.join(quotation_dir, f"{quotation_id}.pdf")

    # If the quotation with the give ID doesn't exist in the exported quotations table, add it
    

    # Change the status of the quotation to approved
    conn.execute(
        text(f"UPDATE exported_quotations SET status = 'Approved' WHERE quotation_id = {quotation_id}")
    )
    conn.commit()
    


    # Add QR code to the Excel
    qr_code = convert_url_to_qr_code(f"http://localhost:5000/view?quotation_id={quotation_id}")
    path = f"{quotation_dir}/{quotation_id}.png"
    qr_code = qr_code.save(path, "PNG")
    img = openpyxl.drawing.image.Image(path)
    img.anchor = "C24"
    img.height = 100
    img.width = 100
    sheet.add_image(img)
    
    # If the excel file doesn't exist, make it and save it
    if not os.path.exists(quotation_file_path):
        #Workesheet object has no attribute save
        quotations.save(filename=quotation_file_path)


    # Open Microsoft Excel
    initialize_com()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    quotation_file_path = os.path.abspath(quotation_file_path)
    pdf_file_path = os.path.abspath(pdf_file_path)
    sheets = excel.Workbooks.Open(quotation_file_path)
    # Get all worksheets
    work_sheets = sheets.Worksheets

    # Convert into PDF File
    # Merge all the worksheets into one PDF
    sheets.ExportAsFixedFormat(0, pdf_file_path)

"""

def submit_quotation_to_db(employee_id: int, quotations, sheet, sent_to_approve = False) -> bool:
    no_of_quotations = conn.execute(
        text("SELECT MAX(quotation_id) FROM exported_quotations")
    ).all()[0][0]
    if not no_of_quotations:
        no_of_quotations = 0
    if sent_to_approve == False:
        quotation_id = no_of_quotations + 1
    else:
        quotation_id = no_of_quotations 
    submission_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    quotation_url = f"http://localhost:5000/view?quotation_id={quotation_id}"  # TODO: Change this to the actual URL
    quotation_file_path = os.path.join(excel_quotation_dir, f"{quotation_id}.xlsx")
    pdf_file_path = os.path.join(quotation_dir, f"{quotation_id}.pdf")
    # Check access level of the user, if the user is an admin, set the status to approved, else set it to pending
    if session["access_level"] == "Administrator" or session["access_level"] == "Developer":
        status = "Approved"
    else:
        status = "Pending"
    if not sent_to_approve:
        conn.execute(
            text(
                f"INSERT INTO exported_quotations(quotation_id, submission_date, employee_id, quotation_url, quotation_file_path, status) VALUES ('{quotation_id}','{submission_date}', {employee_id}, '{quotation_url}', '{quotation_file_path}', '{status}')"
            )
        )
        conn.commit()
    

    # Save the quotation to the file system
    qr_code = convert_url_to_qr_code(quotation_url)
    path = f"{quotation_dir}/{quotation_id}.png"
    qr_code = qr_code.save(path, "PNG")
    print(type(qr_code))
    print(path)
    # Add QR code to the Excel

    img = openpyxl.drawing.image.Image(path)
    img.anchor = "C24"
    img.height = 100
    img.width = 100
    sheet.add_image(img)

    # Save the quotation to the file system
    quotations.save(filename=quotation_file_path)

    # Open Microsoft Excel

    quotation_file_path = os.path.abspath(quotation_file_path)
    pdf_file_path = os.path.abspath(pdf_file_path)

    initialize_com()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    sheets = excel.Workbooks.Open(quotation_file_path)
    # Get all worksheets
    work_sheets = sheets.Worksheets

    # Convert into PDF File
    # Merge all the worksheets into one PDF
    sheets.ExportAsFixedFormat(0, pdf_file_path)
    # Close the Excel application
    excel.Quit()
    return pdf_file_path


def update_password(user_name: str, new_password) -> bool:
    # Make a SQL request to the DB to update the password hash for the user
    new_hash = generate_password_hash(new_password, method="pbkdf2", salt_length=16)

    query = f"UPDATE authorized_personnel SET Password = '{new_hash}' WHERE Employee_Name = '{user_name}'"
    response = conn.execute(text(query))
    conn.commit()
    return (
        response.rowcount == 1
    )  # If the row was updated, return True, else return False


# Return the image of the QR code to the calling function
def convert_url_to_qr_code(url: str, rounded_corners=True, logo_path=None) -> PIL.Image:
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_L)
    qr.add_data(url)

    if rounded_corners and logo_path:
        img = qr.make_image(
            image_factory=StyledPilImage,
            module_drawer=RoundedModuleDrawer(),
            embeded_image_path=logo_path,
        )
    elif rounded_corners:
        img = qr.make_image(
            image_factory=StyledPilImage, module_drawer=RoundedModuleDrawer()
        )
    elif logo_path:
        img = qr.make_image(image_factory=StyledPilImage, embeded_image_path=logo_path)
    else:
        img = qr.make_image()
    print(type(img))
    return img



"""
DONE: Modify the database schema:

Add a new table for pending quotations
Add a status field to the existing quotations table (e.g., 'pending', 'approved', 'rejected')


DONE: Update the quotation creation process:

When a non-authorized user creates a quotation, save it to the pending quotations table instead of directly submitting it
Assign a 'pending' status to these quotations


DONE: Create a pending queue interface:

Develop a new page or section for authorized users to view pending quotations
This page should list all quotations with 'pending' status


DONE: Implement approval/rejection functionality:

Add buttons or controls for authorized users to approve or reject pending quotations
When approved, move the quotation from the pending table to the main quotations table and update its status
If rejected, either delete the pending quotation or mark it as rejected


DONE: Modify user permissions:

Ensure that only authorized users can access the pending queue and approve/reject quotations
Update your access control logic to reflect these new permissions


DONE: Update the quotation viewing process:

Modify the existing quotation view to show the status of each quotation
Potentially hide or visually distinguish pending quotations from approved ones


Notification system (optional):

Implement a way to notify authorized users when new pending quotations are created
Notify the creating user when their quotation is approved or rejected


Reporting and analytics (optional):

Add functionality to track metrics like approval rates, time to approval, etc.
This could be used as a data science project of sorts. Who approves what and in how long. On average, etc.

TODO: Do not allow spaces for usernames, make sure you do that for both frontend and backend validation

DONE: Make all prices editable within Quotation creation page

DONE: Admin can approve or reject price offer with a single click instead of editing prices manually

DONE: If the submit button is clicked twice it crashes the site. Work on fixing that JS That disables button on click!

DONE: Price edits don't work now? For some reason?

TODO: Add a search quotations option that allows you to download a quotation by entering the quotation ID or search by date

TODO: Make the search categories dropdowns instead of text fields.

TODO: Add a search option from date to date

"""